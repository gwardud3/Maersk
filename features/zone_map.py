import streamlit as st
import geopandas as gpd
import matplotlib.pyplot as plt
import pandas as pd
import matplotlib.patches as mpatches
import os
from openpyxl import load_workbook
from io import BytesIO
from copy import copy

# ---------------- Resource path (repo-root safe) ----------------
def resource_path(relative_path: str) -> str:
    return os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..",
        relative_path
    )

# ---------------- Cached ZIP3 shapes (GPKG) ----------------
@st.cache_resource
def load_zip3_shapes():
    gdf = gpd.read_file(
        resource_path("shapefiles/zip3_simplified.gpkg"),
        engine="fiona"  # more stable on Streamlit Cloud
    )
    gdf["zip3"] = gdf["zip3"].astype(str).str.zfill(3)
    return gdf

# ---------------- Heavy processing function ----------------
def process_data(origin_list, customer_name):
    progress_text = st.empty()

    # Step 1: Load Excel
    progress_text.info("Loading Excel file...")
    excel_path = resource_path("Maersk Zones.xlsx")
    MasterZone_df = pd.read_excel(excel_path)

    # Step 2: Process Data
    progress_text.info("Processing zone data...")
    MasterZone_df["OriginZip"] = MasterZone_df["Set_ID"].astype(str).str.zfill(3)
    MasterZone_df["DestZipMin"] = MasterZone_df["Min_Zip_Int"].astype(int)
    MasterZone_df["DestZipMax"] = MasterZone_df["Max_Zip_Int"].astype(int)
    MasterZone_df["Zone"] = MasterZone_df["Zone"].astype(int)

    filtered = MasterZone_df[
        MasterZone_df["OriginZip"].isin(origin_list)
    ].copy()

    filtered["DestZipRange"] = filtered.apply(
        lambda r: range(r.DestZipMin, r.DestZipMax + 1),
        axis=1
    )

    expanded_df = filtered.explode("DestZipRange")
    expanded_df["zip3"] = expanded_df["DestZipRange"].astype(str).str.zfill(3)
    expanded_df = expanded_df[["zip3", "Zone", "OriginZip"]]

    min_zones = expanded_df.groupby("zip3")["Zone"].min().reset_index()
    min_zone_df = expanded_df.merge(min_zones, on=["zip3", "Zone"])

    expanded_df = min_zone_df.copy()

    # Step 3: Load ZIP3 shapes
    progress_text.info("Loading ZIP3 map shapes...")
    zip3_shapes = load_zip3_shapes()

    expanded_df["zip3"] = expanded_df["zip3"].astype(str).str.zfill(3)
    zip3_shapes = zip3_shapes.merge(expanded_df, on="zip3", how="left")

    # Step 4: Plot
    progress_text.info("Rendering map...")
    zone_colors = {
        1: "#001624", 2: "#00243D", 3: "#004A73",
        4: "#0073AB", 5: "#2392BE", 6: "#42B0D5",
        7: "#72C8E3", 8: "#A1D8EF", 9: "#B5E0F5"
    }

    fig, ax = plt.subplots(figsize=(15, 10))

    # State boundaries (lightweight, OK to reload)
    states = gpd.read_file(
        resource_path("shapefiles/states_preprocessed.gpkg"),
        engine="fiona"
    )
    states.boundary.plot(ax=ax, linewidth=0.5, edgecolor="black")

    ax.set_facecolor("#a3a3a3")

    zip3_plot_colors = zip3_shapes["Zone"].map(zone_colors).fillna("#CCCCCC")
    zip3_shapes.plot(ax=ax, color=zip3_plot_colors, linewidth=0)

    # Continental US view
    ax.set_xlim(-130, -65)
    ax.set_ylim(24, 50)
    ax.set_aspect(1.2, adjustable="box")

    used_zones = sorted(
        z for z in zip3_shapes["Zone"].dropna().unique()
        if z != 9
    )

    legend_handles = [
        mpatches.Patch(color=zone_colors[z], label=str(z))
        for z in used_zones
    ]
    
    ax.legend(
        handles=legend_handles,
        title="Zone",
        loc="lower left",
        fontsize="small"
    )

    ax.set_title(f"Zone Map – {customer_name}", fontsize=16)
    ax.axis("off")
    plt.tight_layout()

    progress_text.success("Done!")

    return fig, expanded_df


def process_export_data(origin_list):
    # Load Excel
    excel_path = resource_path("Maersk Zones.xlsx")
    df = pd.read_excel(excel_path)

    # Clean columns
    df["OriginZip"] = df["Set_ID"].astype(str).str.zfill(3)
    df["DestZipMin"] = df["Min_Zip_Int"].astype(int)
    df["DestZipMax"] = df["Max_Zip_Int"].astype(int)
    df["Zone"] = df["Zone"].astype(int)

    # Filter only selected origins
    df = df[df["OriginZip"].isin(origin_list)].copy()

    # Expand ZIP ranges → 1:1 mapping
    df["DestZipRange"] = df.apply(
        lambda r: range(r["DestZipMin"], r["DestZipMax"] + 1),
        axis=1
    )

    df = df.explode("DestZipRange")

    # Convert to ZIP3
    df["zip3"] = df["DestZipRange"].astype(str).str.zfill(3)

    # Keep only what we need
    df = df[["zip3", "OriginZip", "Zone"]]

    return df
# ---------------- Streamlit Feature Entry Point ----------------
def zone_map_app():
    st.header("📦 Zone Map Generator")

    origin_input = st.text_input(
        "Enter 3-Digit Origin ZIPs (comma separated)"
    )
    customer_name = st.text_input("Customer Name")

    if st.button("Generate Map"):
        origin_list = [
            o.strip().zfill(3)
            for o in origin_input.split(",")
            if o.strip().isdigit() and len(o.strip()) <= 3
        ]

        if not origin_list:
            st.error("Please enter at least one valid 3-digit Origin ZIP.")
            return

        if not customer_name:
            st.error("Please enter a Customer Name.")
            return

        with st.spinner("Processing… this may take a moment"):
            fig, expanded_df = process_data(origin_list, customer_name)

        st.pyplot(fig)

        # ================================
        # 📥 EXPORT USING TEMPLATE (MATRIX FORMAT)
        # ================================
        
        export_df = process_export_data(origin_list)
        template_path = resource_path("assets/ZoningTemplate.xlsx")
        
        wb = load_workbook(template_path)
        ws = wb.active  # or specify sheet
        
        # ================================
        # 🧠 BUILD MATRIX (ZIP3 x ORIGIN)
        # ================================
        
        # Pivot: rows = zip3, columns = OriginZip, values = Zone
        pivot_df = export_df.pivot_table(
            index="zip3",
            columns="OriginZip",
            values="Zone",
            aggfunc="first"
        ).reset_index()
        
        # Ensure ALL user-selected origins appear as columns
        for origin in origin_list:
            if origin not in pivot_df.columns:
                pivot_df[origin] = None
        
        # Reorder columns: zip3 first, then origins in user order
        pivot_df = pivot_df[["zip3"] + origin_list]
        
        # ================================
        # 🎨 COPY TEMPLATE FORMATTING
        # ================================
        
        def copy_column_format(ws, source_col, target_col):
            for row in range(4, data_start_row + len(pivot_df) + 5):
                source_cell = ws.cell(row=row, column=source_col)
                target_cell = ws.cell(row=row, column=target_col)
        
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)
        
        # Template starts:
        header_row = 4
        data_start_row = 5
        origin_start_col = 2  # Column B
        
        existing_origin_cols = ws.max_column - (origin_start_col - 1)
        
        # Expand formatting if needed
        if len(origin_list) > existing_origin_cols:
            for i in range(existing_origin_cols, len(origin_list)):
                source_col = origin_start_col + existing_origin_cols - 1
                target_col = origin_start_col + i
                copy_column_format(ws, source_col, target_col)
        
        # ================================
        # 🧹 CLEAR OLD DATA
        # ================================
        
        for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        
        # ================================
        # ✍️ WRITE HEADERS (ROW 4)
        # ================================
        
        # Column A header stays as-is
        
        for col_idx, origin in enumerate(origin_list, start=origin_start_col):
            ws.cell(row=header_row, column=col_idx, value=origin)
        
        # ================================
        # ✍️ WRITE DATA (ROW 5+)
        # ================================
        
        for r_idx, row in enumerate(pivot_df.itertuples(index=False), start=data_start_row):
            
            # Column A = zip3
            ws.cell(row=r_idx, column=1, value=row[0])
        
            # Columns B+ = zones
            for c_idx, value in enumerate(row[1:], start=origin_start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # ================================
        # 💾 SAVE FOR DOWNLOAD
        # ================================
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        st.download_button(
            label="📥 Download Zoning Table (Template)",
            data=output,
            file_name=f"{customer_name}_zoning_table.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
